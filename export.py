from flask import send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
from models import Registro
from config import aplicar_filtros
from datetime import datetime

def exportar_registros_para_excel(filtro_posto, filtro_data_html, filtro_coleta):
    """
    Gera um arquivo XLSX com os registros filtrados.
    """
    
    # 1. Aplicar Filtros (a função já aplica a ordenação correta)
    query = Registro.query
    query = aplicar_filtros(query, filtro_posto, filtro_data_html, filtro_coleta)
    registros = query.all()
    
    # 2. Criar o Workbook e a planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Registros de Procedimentos"

    # Estilos
    header_fill = PatternFill(start_color="337ab7", end_color="337ab7", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    center_aligned_text = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_aligned_text = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 3. Cabeçalho
    cabecalho = [
        "Posto", "Data (Dia/Mês/Ano)", "Hora Início", "Hora Término", 
        "Mesa/Local", "Coleta (SIM/NÃO)", "Destino Retaguarda", "Procedimento Completo"
    ]
    ws.append(cabecalho)
    
    # Aplicar estilos ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_aligned_text

    # Definir larguras das colunas
    ws.column_dimensions['A'].width = 20 
    ws.column_dimensions['B'].width = 18 
    ws.column_dimensions['C'].width = 12 
    ws.column_dimensions['D'].width = 12 
    ws.column_dimensions['E'].width = 15 
    ws.column_dimensions['F'].width = 10 
    ws.column_dimensions['G'].width = 20 
    ws.column_dimensions['H'].width = 70 

    # 4. Inserir Dados
    for registro in registros:
        
        coleta_valor = registro.retaguarda_sim_nao
        destino = registro.retaguarda_destino if registro.retaguarda_sim_nao == 'SIM' else ""
        
        linha = [
            registro.posto,
            registro.data,
            registro.hora_inicio,
            registro.hora_termino,
            registro.numero_mesa,
            coleta_valor,
            destino,
            registro.procedimento
        ]
        ws.append(linha)
        
        # Aplicar bordas e alinhamento
        for cell in ws[ws.max_row]:
            cell.border = thin_border
            if cell.column_letter in ['H']:
                cell.alignment = left_aligned_text
            else:
                cell.alignment = center_aligned_text

    # 5. Salvar em memória
    data = BytesIO()
    wb.save(data)
    data.seek(0)
    
    # 6. Retornar como arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"Registros_PPT_{timestamp}.xlsx"
    
    return send_file(
        data,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
