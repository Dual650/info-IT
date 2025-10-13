from flask import Flask, render_template, request, redirect, url_for, send_file, jsonify
from datetime import datetime
from flask_sqlalchemy import SQLAlchemy
import os
import io
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import Workbook
import io

# --- Opções de Postos de Trabalho ---
POSTOS = [
    "Andradina", "Assis", "Avaré", "Bauru", "Birigui", "Botucatu",
    "Dracena", "Itapetininga", "Itu", "Jahu", "Marília",
    "Ourinhos", "Penápolis", "Presidente Prudente", "Tatuí", "Tupã"
]

app = Flask(__name__)

# --- Configuração do Banco de Dados (Pronto para Render/PostgreSQL) ---
DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///site.db').replace("postgres://", "postgresql://", 1)
app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

# --- Definição do Modelo (Tabela) do Banco de Dados ---
class Registro(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    posto = db.Column(db.String(50), nullable=False)
    computador_coleta = db.Column(db.String(3), nullable=False) 
    data = db.Column(db.String(10), nullable=False) 
    hora_inicio = db.Column(db.String(5), nullable=False)
    hora_termino = db.Column(db.String(5), nullable=False)
    procedimento = db.Column(db.Text, nullable=False)
    timestamp_registro = db.Column(db.DateTime, default=datetime.utcnow) 

    def __repr__(self):
        return f"Registro('{self.posto}', '{self.data}', '{self.hora_inicio}')"

# Cria as tabelas do banco de dados
with app.app_context():
    db.create_all()

# --- CONSTANTE para Resumo ---
RESUMO_MAX_CARACTERES = 70 # Define o limite de caracteres para o resumo na tela

# Função auxiliar para aplicar filtros
def aplicar_filtros(query, filtro_posto, filtro_data_html):
    if filtro_posto and filtro_posto != 'Todos':
        query = query.filter(Registro.posto == filtro_posto)
            
    if filtro_data_html:
        try:
            data_obj = datetime.strptime(filtro_data_html, '%Y-%m-%d')
            data_formatada = data_obj.strftime('%d/%m/%Y')
            query = query.filter(Registro.data == data_formatada)
        except ValueError:
            pass
            
    return query.order_by(Registro.timestamp_registro.desc())

# --- Rota 1: Registro de Novo Procedimento (Sem Alterações) ---
@app.route('/', methods=['GET', 'POST'])
def formulario_registro():
    data_de_hoje = datetime.now().strftime('%d/%m/%Y')
    
    if request.method == 'POST':
        posto = request.form.get('posto')
        computador_coleta = request.form.get('computador_coleta')
        hora_inicio = request.form.get('hora_inicio')
        hora_termino = request.form.get('hora_termino')
        procedimento = request.form.get('procedimento')
        
        novo_registro = Registro(
            posto=posto,
            computador_coleta=computador_coleta,
            data=data_de_hoje,
            hora_inicio=hora_inicio,
            hora_termino=hora_termino,
            procedimento=procedimento
        )
        db.session.add(novo_registro)
        db.session.commit()

        return redirect(url_for('formulario_registro'))

    return render_template(
        'index.html', 
        postos=POSTOS, 
        data_de_hoje=data_de_hoje
    )

# --- Rota 2: Consulta de Registros Antigos (Apenas exibe a página HTML) ---
@app.route('/consultar', methods=['GET'])
def consultar_registro():
    """
    Simplesmente renderiza a página de consulta. O JS fará a primeira busca.
    """
    filtro_posto = request.args.get('posto')
    filtro_data_html = request.args.get('data')

    return render_template(
        'consultar.html', 
        postos=POSTOS, 
        filtro_posto=filtro_posto or 'Todos', # Garante que o dropdown inicie com o valor correto
        filtro_data=filtro_data_html
    )

# --- Rota: Retorna os dados em JSON para o JavaScript ---
@app.route('/registros_json', methods=['GET'])
def registros_json():
    """
    Retorna os registros filtrados em formato JSON, com o procedimento resumido.
    """
    filtro_posto = request.args.get('posto')
    filtro_data_html = request.args.get('data')
    
    query = Registro.query
    query = aplicar_filtros(query, filtro_posto, filtro_data_html)
    registros = query.all()

    # Formata os dados para JSON, aplicando o resumo
    registros_formatados = []
    for r in registros:
        
        # Lógica de Resumo do Procedimento
        procedimento_completo = r.procedimento
        procedimento_resumo = procedimento_completo
        if len(procedimento_completo) > RESUMO_MAX_CARACTERES:
            procedimento_resumo = procedimento_completo[:RESUMO_MAX_CARACTERES] + '...'
            
        registros_formatados.append({
            'posto': r.posto,
            'computador_coleta': r.computador_coleta,
            'data': r.data,
            'hora_inicio': r.hora_inicio,
            'hora_termino': r.hora_termino,
            # Campo que será exibido na tela
            'procedimento_resumo': procedimento_resumo, 
            # Campo que será usado para exportação (não usado no front-end, mas importante saber que existe)
            'procedimento_completo': procedimento_completo,
            'id': r.id
        })

    return jsonify(registros_formatados)


# --- Rota 3: Exportar para XLSX (Corrigida a conversão de tipo para Hora) ---
@app.route('/exportar', methods=['GET'])
def exportar_registros():
    filtro_posto = request.args.get('posto')
    filtro_data_html = request.args.get('data')

    query = Registro.query
    query = aplicar_filtros(query, filtro_posto, filtro_data_html)
    registros = query.all()

    if not registros:
        # Se não houver registros, retorna para a página de consulta com uma mensagem
        return redirect(url_for('consultar_registro', erro='Nenhum registro encontrado para exportar.'))

    # 1. Preparar os dados para o DataFrame (USA PROCEDIMENTO COMPLETO)
    dados = []
    for r in registros:
        try:
            # Converte a data de DD/MM/YYYY para um objeto date (ideal para Pandas/Excel)
            data_obj = datetime.strptime(r.data, '%d/%m/%Y').date()
        except ValueError:
            data_obj = r.data
            
        # --- CORREÇÃO APLICADA AQUI: Converte strings de hora para objetos time ---
        try:
            # Converte a hora de "HH:MM" (string) para objeto time (essencial para o Excel)
            inicio_obj = datetime.strptime(r.hora_inicio, '%H:%M').time()
        except ValueError:
            inicio_obj = r.hora_inicio # Mantém como string se a conversão falhar
            
        try:
            termino_obj = datetime.strptime(r.hora_termino, '%H:%M').time()
        except ValueError:
            termino_obj = r.hora_termino # Mantém como string se a conversão falhar
        # --------------------------------------------------------------------------

        dados.append({
            'Posto': r.posto,
            'Computador da coleta?': r.computador_coleta,  
            'Data': data_obj,
            'Início': inicio_obj, # AGORA É UM OBJETO TIME
            'Término': termino_obj, # AGORA É UM OBJETO TIME
            'Procedimento Realizado': r.procedimento # <--- TEXTO ORIGINAL AQUI
        })

    df = pd.DataFrame(dados)

    # 2. Configurar o Writer e o Workbook
    output = io.BytesIO()
    # Fecha o writer ao final do bloco 'with'
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Registros Técnicos')
        workbook = writer.book
        sheet = writer.sheets['Registros Técnicos']

        # --- 3. Definir Estilos ---
        FILL_GRAY = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        header_font = Font(size=16, bold=True, color="000000")
        header_alignment = Alignment(horizontal='center', vertical='center')  
        
        data_font = Font(size=12, color="000000")  
        data_font_bold = Font(size=12, bold=True, color="000000")
        data_alignment = Alignment(horizontal='center', vertical='top')
        procedimento_alignment = Alignment(horizontal='left', vertical='top', wrapText=True)

        # --- 4. Aplicar Estilos de Cabeçalho e Largura de Coluna ---
        
        col_config = {
            'Posto': 15,
            'Computador da coleta?': 18,  
            'Data': 15,
            'Início': 12,
            'Término': 12,
            'Procedimento Realizado': 60
        }
        
        for i, col_name in enumerate(df.columns):
            col_letter = sheet.cell(row=1, column=i+1).column_letter
            
            sheet.column_dimensions[col_letter].width = col_config.get(col_name, 15)
            
            header_cell = sheet.cell(row=1, column=i+1)
            header_cell.font = header_font
            header_cell.fill = FILL_GRAY
            header_cell.alignment = header_alignment

        # --- 5. Aplicar Estilos e Formatos ao Corpo da Planilha (Dados) ---
        
        DEFAULT_ROW_HEIGHT = 15
        sheet.default_row_dimension.height = DEFAULT_ROW_HEIGHT  

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_col=len(df.columns))):
            
            valor_coleta = row[1].value  
            
            sheet.row_dimensions[row_idx + 2].height = DEFAULT_ROW_HEIGHT  

            for col_idx, cell in enumerate(row):
                col_name = df.columns[col_idx]
                
                cell.alignment = data_alignment  
                cell.font = data_font  
                
                if col_name == 'Computador da coleta?':
                    cell.font = data_font_bold
                    if valor_coleta == 'Sim':
                        cell.fill = FILL_GREEN
                    elif valor_coleta == 'Não':
                        cell.fill = FILL_RED
                
                elif col_name == 'Data':
                    cell.number_format = 'DD/MM/YYYY'  
                
                elif col_name in ['Início', 'Término']:
                    # Esta linha agora funciona, pois o dado na célula é um objeto time
                    cell.number_format = 'HH:MM'
                
                elif col_name == 'Procedimento Realizado':
                    cell.alignment = procedimento_alignment  
                    sheet.row_dimensions[row_idx + 2].height = 40  


    # 6. Salvar e Retornar
    output.seek(0)

    data_exportacao = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"registros_tecnicos_{data_exportacao}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        download_name=filename,
        as_attachment=True
    )

if __name__ == '__main__':
    app.run(debug=True)
